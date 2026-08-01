from io import BytesIO
from typing import ClassVar

from openpyxl import load_workbook

from syncmark_demo.domain.models import RawItem


class OpenpyxlReader:
    REQUIRED_HEADERS: ClassVar[set[str]] = {"GTIN", "category", "quantity", "product_name"}

    def read(self, content: bytes) -> list[RawItem]:
        try:
            sheet = load_workbook(BytesIO(content), read_only=True, data_only=True).active
        except Exception as error:
            raise ValueError("invalid_xlsx") from error
        if sheet is None:
            raise ValueError("empty_xlsx")
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError("empty_xlsx")
        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        if set(headers) != self.REQUIRED_HEADERS:
            raise ValueError("invalid_headers")
        positions = {name: headers.index(name) for name in headers}
        return [
            RawItem(index, *[None if values[positions[name]] is None else str(values[positions[name]]) for name in ("GTIN", "category", "quantity", "product_name")])
            for index, values in enumerate(rows[1:], start=2)
            if any(value is not None and str(value).strip() for value in values)
        ]
